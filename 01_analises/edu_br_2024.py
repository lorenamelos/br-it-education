"""
ANÁLISE DO TRIPLO DÉFICIT TECNOLÓGICO NA EDUCAÇÃO BRASILEIRA
Dados: TIC Educação 2023 (Escolas) + TIC Educação 2024 (Escolas e Alunos)
Autor: Análise Educacional BR
Data: Janeiro 2025
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from pathlib import Path

# Configurações
plt.style.use('seaborn-v0_8-darkgrid')
sns.set_palette("husl")
plt.rcParams['figure.figsize'] = (16, 10)
plt.rcParams['font.size'] = 10

# ============================================================================
# PARTE 1: ANÁLISE DE INFRAESTRUTURA DAS ESCOLAS (2023 e 2024)
# ============================================================================

def listar_abas_disponiveis(arquivo_path):
    """
    Lista todas as abas disponíveis no arquivo Excel
    Útil para debug quando os nomes das abas mudam
    """
    import openpyxl
    
    wb = openpyxl.load_workbook(arquivo_path, read_only=True, data_only=True)
    abas = wb.sheetnames
    wb.close()
    
    print(f"\n📋 Abas disponíveis em {Path(arquivo_path).name}:")
    for i, aba in enumerate(abas, 1):
        print(f"  {i:2}. {aba}")
    print(f"\n  Total: {len(abas)} abas")
    
    return abas

def encontrar_aba(arquivo_path, nome_base):
    """
    Encontra o nome correto da aba, considerando variações como A3, A3_1, etc.
    """
    import openpyxl
    
    # Abrir workbook para listar abas
    wb = openpyxl.load_workbook(arquivo_path, read_only=True, data_only=True)
    abas_disponiveis = wb.sheetnames
    wb.close()
    
    # Procurar pela aba exata primeiro
    if nome_base in abas_disponiveis:
        return nome_base
    
    # Procurar por variações com underscore
    for aba in abas_disponiveis:
        # Verificar se começa com o nome base seguido de underscore
        if aba.startswith(f"{nome_base}_"):
            print(f"  ℹ️  Usando aba '{aba}' para '{nome_base}'")
            return aba
    
    # Se não encontrar, retornar None
    return None

def carregar_dados_escolas_2023(arquivo_path, aba='B4A'):
    """
    Carrega dados de infraestrutura das escolas (TIC 2023)
    Aba B4A: Número de alunos por computador desktop
    """
    print(f"\n📂 Carregando {aba} do TIC Educação 2023...")
    
    df = pd.read_excel(arquivo_path, sheet_name=aba, header=None)
    
    # Identificar linha de cabeçalho (geralmente linha 2)
    header_row = 2
    df.columns = df.iloc[header_row]
    df = df.iloc[header_row+1:].reset_index(drop=True)
    
    # Renomear colunas
    df.columns = ['Categoria', 'Subcategoria', 'Sem_computador', 'Ate_5_alunos',
                  'De_5_1_a_10', 'De_10_1_a_15', 'De_15_1_a_20', 'De_20_1_a_30',
                  'De_30_1_a_40', 'De_40_1_a_50', 'De_50_1_a_100', 'Mais_de_100']
    
    # Converter para numérico
    cols_numericas = df.columns[2:]
    for col in cols_numericas:
        df[col] = pd.to_numeric(df[col], errors='coerce')
    
    # Remover linhas de cabeçalhos intermediários
    df = df[df['Sem_computador'].notna()]
    
    return df

def encontrar_aba(arquivo_path, nome_base):
    """
    Encontra o nome correto da aba, considerando variações como A3, A3_1, etc.
    """
    import openpyxl
    
    # Abrir workbook para listar abas
    wb = openpyxl.load_workbook(arquivo_path, read_only=True, data_only=True)
    abas_disponiveis = wb.sheetnames
    wb.close()
    
    # Procurar pela aba exata primeiro
    if nome_base in abas_disponiveis:
        return nome_base
    
    # Procurar por variações com underscore
    for aba in abas_disponiveis:
        # Verificar se começa com o nome base seguido de underscore
        if aba.startswith(f"{nome_base}_"):
            print(f"  ℹ️  Usando aba '{aba}' para '{nome_base}'")
            return aba
    
    # Se não encontrar, retornar None
    return None

def carregar_dados_escolas_2024(arquivo_path):
    """
    Carrega dados atualizados de escolas TIC 2024
    Processa abas: A8 (acesso), A3 (velocidade), J1 (formação)
    Lida com variações nos nomes das abas (A3, A3_1, etc.)
    """
    print(f"\n📂 Carregando dados TIC Educação 2024 - Escolas...")
    
    dados = {}
    
    # A8 - Acesso a computador + internet
    aba_a8 = encontrar_aba(arquivo_path, 'A8')
    if aba_a8 is None:
        raise ValueError("❌ Aba A8 não encontrada no arquivo!")
    
    df_a8 = pd.read_excel(arquivo_path, sheet_name=aba_a8, header=None)
    df_a8.columns = df_a8.iloc[2]
    df_a8 = df_a8.iloc[3:].reset_index(drop=True)
    dados['A8'] = df_a8
    print(f"  ✅ A8: {len(df_a8)} linhas carregadas")
    
    # A3 - Velocidade da internet
    aba_a3 = encontrar_aba(arquivo_path, 'A3')
    if aba_a3 is None:
        raise ValueError("❌ Aba A3 (ou variação) não encontrada no arquivo!")
    
    df_a3 = pd.read_excel(arquivo_path, sheet_name=aba_a3, header=None)
    df_a3.columns = df_a3.iloc[2]
    df_a3 = df_a3.iloc[3:].reset_index(drop=True)
    dados['A3'] = df_a3
    print(f"  ✅ A3: {len(df_a3)} linhas carregadas")
    
    # J1 - Formação docente
    aba_j1 = encontrar_aba(arquivo_path, 'J1')
    if aba_j1 is None:
        raise ValueError("❌ Aba J1 (ou variação) não encontrada no arquivo!")
    
    df_j1 = pd.read_excel(arquivo_path, sheet_name=aba_j1, header=None)
    df_j1.columns = df_j1.iloc[2]
    df_j1 = df_j1.iloc[3:].reset_index(drop=True)
    dados['J1'] = df_j1
    print(f"  ✅ J1: {len(df_j1)} linhas carregadas")
    
    print("✅ Dados de escolas 2024 carregados com sucesso!")
    
    return dados

def carregar_dados_alunos_2024(arquivo_path):
    """
    Carrega dados de uso de IA por alunos (TIC 2024)
    Aba G6: Recursos digitais em pesquisas escolares
    """
    print(f"\n📂 Carregando dados TIC Educação 2024 - Alunos (IA)...")
    
    df = pd.read_excel(arquivo_path, sheet_name='G6', header=None)
    
    # Dados começam na linha 4 (índice 3)
    # Coluna 32 = "Sim" para uso de IA
    # Coluna 33 = "Não" para uso de IA
    
    dados_ia = {
        'brasil': {
            'usa_ia': df.iloc[3, 32],  # Linha TOTAL
            'nao_usa_ia': df.iloc[3, 33],
            'percentual': None  # Calcular depois
        },
        'regioes': {},
        'etapas': {}
    }
    
    # Calcular percentual Brasil
    total = dados_ia['brasil']['usa_ia'] + dados_ia['brasil']['nao_usa_ia']
    dados_ia['brasil']['percentual'] = (dados_ia['brasil']['usa_ia'] / total) * 100
    
    print(f"✅ Dados de uso de IA extraídos: {dados_ia['brasil']['percentual']:.1f}% usam IA")
    
    return dados_ia

# ============================================================================
# PARTE 2: ANÁLISE DO TRIPLO DÉFICIT
# ============================================================================

def calcular_indice_prontidao(dados_escolas_2024):
    """
    Calcula o Índice de Prontidão para IA baseado em 3 pilares:
    1. Acesso a computador + internet (A8)
    2. Qualidade da conexão ≥3 Mbps (A3)
    3. Formação docente (J1)
    """
    print("\n" + "="*80)
    print("CALCULANDO ÍNDICE DE PRONTIDÃO PARA IA GENERATIVA")
    print("="*80)
    
    # Extrair dados do Brasil (primeira linha de dados)
    a8_brasil = dados_escolas_2024['A8'].iloc[0]
    a3_brasil = dados_escolas_2024['A3'].iloc[0]
    j1_brasil = dados_escolas_2024['J1'].iloc[0]
    
    # Pilar 1: % com acesso (computador + internet)
    com_acesso = pd.to_numeric(a8_brasil.iloc[2], errors='coerce')
    sem_acesso = pd.to_numeric(a8_brasil.iloc[3], errors='coerce')
    total = com_acesso + sem_acesso
    pct_com_acesso = (com_acesso / total) * 100
    
    # Pilar 2: % com conexão adequada (≥3 Mbps)
    # Colunas de velocidade: até 999Kbps, 1-2Mbps, 3-10Mbps, 11-20Mbps, >20Mbps
    conexao_ruim = pd.to_numeric(a3_brasil.iloc[2], errors='coerce') + pd.to_numeric(a3_brasil.iloc[3], errors='coerce')
    conexao_boa = sum([pd.to_numeric(a3_brasil.iloc[i], errors='coerce') for i in range(4, 9)])
    total_internet = conexao_ruim + conexao_boa
    pct_conexao_boa = (conexao_boa / total_internet) * 100
    
    # Pilar 3: % com formação docente
    com_formacao = pd.to_numeric(j1_brasil.iloc[2], errors='coerce')
    sem_formacao = pd.to_numeric(j1_brasil.iloc[3], errors='coerce')
    total_escolas = com_formacao + sem_formacao
    pct_com_formacao = (com_formacao / total_escolas) * 100
    
    # Índice de Prontidão = multiplicação dos 3 pilares
    indice_prontidao = (pct_com_acesso / 100) * (pct_conexao_boa / 100) * (pct_com_formacao / 100) * 100
    
    print(f"\n📊 PILARES DO ÍNDICE:")
    print(f"  1. Acesso (PC+Internet):  {pct_com_acesso:.1f}%")
    print(f"  2. Conexão adequada:      {pct_conexao_boa:.1f}%")
    print(f"  3. Formação docente:      {pct_com_formacao:.1f}%")
    print(f"\n  🎯 ÍNDICE DE PRONTIDÃO:  {indice_prontidao:.1f}%")
    print(f"     (~{(indice_prontidao/100 * total):.0f} escolas de {total:.0f})")
    
    return {
        'pct_acesso': pct_com_acesso,
        'pct_conexao': pct_conexao_boa,
        'pct_formacao': pct_com_formacao,
        'indice': indice_prontidao,
        'escolas_prontas': (indice_prontidao/100 * total),
        'total_escolas': total
    }

def analisar_paradoxo_ia(dados_alunos, indice_prontidao):
    """
    Analisa o paradoxo entre uso estudantil e capacidade escolar
    """
    print("\n" + "="*80)
    print("PARADOXO: USO ESTUDANTIL vs CAPACIDADE ESCOLAR")
    print("="*80)
    
    pct_alunos_usam = dados_alunos['brasil']['percentual']
    pct_escolas_prontas = indice_prontidao['indice']
    
    print(f"\n📊 USO DE IA GENERATIVA:")
    print(f"  • Alunos que USAM IA:        {pct_alunos_usam:.1f}%")
    print(f"  • Escolas PRONTAS para IA:   {pct_escolas_prontas:.1f}%")
    print(f"\n  ⚠️  GAP: {pct_alunos_usam - pct_escolas_prontas:.1f} pontos percentuais")
    print(f"\n💡 INTERPRETAÇÃO:")
    print(f"  Estudantes usam IA APESAR da escola, não POR CAUSA da escola")
    print(f"  Aprendizado acontece de forma autônoma, sem mediação pedagógica")
    
    return {
        'gap': pct_alunos_usam - pct_escolas_prontas,
        'pct_alunos': pct_alunos_usam,
        'pct_escolas': pct_escolas_prontas
    }

# ============================================================================
# PARTE 3: VISUALIZAÇÕES
# ============================================================================

def criar_visualizacoes_completas(indice_prontidao, paradoxo, dados_alunos):
    """
    Cria visualizações consolidadas da análise
    """
    fig = plt.figure(figsize=(18, 12))
    gs = fig.add_gridspec(3, 3, hspace=0.3, wspace=0.3)
    
    # 1. Triplo Déficit (barras)
    ax1 = fig.add_subplot(gs[0, :2])
    pilares = ['Acesso\n(PC+Internet)', 'Conexão\n(≥3 Mbps)', 'Formação\nDocente', 'ÍNDICE\nPRONTIDÃO']
    valores = [
        indice_prontidao['pct_acesso'],
        indice_prontidao['pct_conexao'],
        indice_prontidao['pct_formacao'],
        indice_prontidao['indice']
    ]
    cores = ['#3498db', '#e74c3c', '#f39c12', '#2ecc71']
    
    bars = ax1.bar(pilares, valores, color=cores, edgecolor='black', linewidth=1.5)
    ax1.set_ylabel('Percentual (%)', fontsize=12, fontweight='bold')
    ax1.set_title('Triplo Déficit Tecnológico - Pilares do Índice de Prontidão',
                  fontsize=14, fontweight='bold', pad=20)
    ax1.set_ylim(0, 100)
    ax1.axhline(y=50, color='gray', linestyle='--', alpha=0.5, label='50%')
    ax1.grid(axis='y', alpha=0.3)
    
    # Adicionar valores nas barras
    for bar, valor in zip(bars, valores):
        height = bar.get_height()
        ax1.text(bar.get_x() + bar.get_width()/2., height + 2,
                f'{valor:.1f}%', ha='center', va='bottom',
                fontsize=11, fontweight='bold')
    
    # 2. Paradoxo (comparação)
    ax2 = fig.add_subplot(gs[0, 2])
    categorias = ['Alunos\nusam IA', 'Escolas\nprontas']
    valores_paradoxo = [paradoxo['pct_alunos'], paradoxo['pct_escolas']]
    cores_paradoxo = ['#27ae60', '#e74c3c']
    
    bars2 = ax2.bar(categorias, valores_paradoxo, color=cores_paradoxo,
                    edgecolor='black', linewidth=1.5)
    ax2.set_ylabel('Percentual (%)', fontsize=11, fontweight='bold')
    ax2.set_title('O Paradoxo da IA', fontsize=13, fontweight='bold', pad=15)
    ax2.set_ylim(0, 100)
    ax2.grid(axis='y', alpha=0.3)
    
    for bar, valor in zip(bars2, valores_paradoxo):
        height = bar.get_height()
        ax2.text(bar.get_x() + bar.get_width()/2., height + 2,
                f'{valor:.1f}%', ha='center', va='bottom',
                fontsize=10, fontweight='bold')
    
    # Adicionar anotação do gap
    ax2.annotate(f'GAP:\n{paradoxo["gap"]:.1f}pp',
                xy=(0.5, (valores_paradoxo[0] + valores_paradoxo[1])/2),
                xytext=(1.5, (valores_paradoxo[0] + valores_paradoxo[1])/2),
                ha='left', va='center', fontsize=10,
                bbox=dict(boxstyle='round,pad=0.5', facecolor='yellow', alpha=0.7),
                arrowprops=dict(arrowstyle='->', lw=2, color='red'))
    
    # 3. Funil de Exclusão
    ax3 = fig.add_subplot(gs[1, :])
    
    total = indice_prontidao['total_escolas']
    sem_acesso = total * (1 - indice_prontidao['pct_acesso']/100)
    com_acesso = total * (indice_prontidao['pct_acesso']/100)
    conexao_ruim = com_acesso * (1 - indice_prontidao['pct_conexao']/100)
    infra_ok = com_acesso * (indice_prontidao['pct_conexao']/100)
    sem_formacao = infra_ok * (1 - indice_prontidao['pct_formacao']/100)
    prontas = indice_prontidao['escolas_prontas']
    
    etapas = ['Total\nEscolas', 'Sem\nAcesso', 'Com\nAcesso', 'Conexão\nRuim',
              'Infra.\nOK', 'Sem\nFormação', 'PRONTAS']
    valores_funil = [total, -sem_acesso, com_acesso, -conexao_ruim,
                     infra_ok, -sem_formacao, prontas]
    cores_funil = ['#3498db', '#e74c3c', '#2ecc71', '#e67e22',
                   '#9b59b6', '#e74c3c', '#27ae60']
    
    x_pos = np.arange(len(etapas))
    bars3 = ax3.bar(x_pos, valores_funil, color=cores_funil,
                    edgecolor='black', linewidth=1.5)
    
    ax3.set_xticks(x_pos)
    ax3.set_xticklabels(etapas, fontsize=10)
    ax3.set_ylabel('Número de Escolas', fontsize=12, fontweight='bold')
    ax3.set_title('Funil de Exclusão - Do Total ao Índice de Prontidão',
                  fontsize=14, fontweight='bold', pad=20)
    ax3.axhline(y=0, color='black', linewidth=2)
    ax3.grid(axis='y', alpha=0.3)
    
    for bar, valor in zip(bars3, valores_funil):
        height = bar.get_height()
        label_pos = height + 1000 if height > 0 else height - 1000
        ax3.text(bar.get_x() + bar.get_width()/2., label_pos,
                f'{abs(valor):,.0f}', ha='center',
                va='bottom' if height > 0 else 'top',
                fontsize=9, fontweight='bold')
    
    # 4. Texto conclusivo
    ax4 = fig.add_subplot(gs[2, :])
    ax4.axis('off')
    
    texto = f"""
    CONCLUSÕES PRINCIPAIS:
    
    • Apenas {indice_prontidao['indice']:.1f}% das escolas brasileiras estão minimamente preparadas para integrar IA generativa
    
    • {paradoxo['pct_alunos']:.1f}% dos alunos usam IA, mas a escola não consegue mediar pedagogicamente esse uso
    
    • TRIPLO DÉFICIT: {100-indice_prontidao['pct_acesso']:.1f}% sem acesso + {100-indice_prontidao['pct_conexao']:.1f}% conexão ruim + {100-indice_prontidao['pct_formacao']:.1f}% sem formação
    
    • Estudantes aprendem IA SOZINHOS, sem orientação crítica ou ética
    
    • Desigualdade regional amplia exclusão digital em múltiplas camadas
    """
    
    ax4.text(0.5, 0.5, texto, transform=ax4.transAxes,
            fontsize=12, verticalalignment='center', horizontalalignment='center',
            bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.5),
            family='monospace')
    
    plt.suptitle('Análise do Triplo Déficit Tecnológico - TIC Educação 2024',
                 fontsize=16, fontweight='bold', y=0.98)
    
    return fig

# ============================================================================
# PARTE 4: EXPORTAÇÃO DE DADOS
# ============================================================================

def exportar_dados_consolidados(indice_prontidao, paradoxo, dados_alunos, output_dir='./resultados'):
    """
    Exporta todos os dados processados em CSV
    """
    Path(output_dir).mkdir(exist_ok=True)
    
    # 1. Resumo do Índice de Prontidão
    df_indice = pd.DataFrame([{
        'Indicador': 'Acesso (PC+Internet)',
        'Percentual': f"{indice_prontidao['pct_acesso']:.1f}%",
        'Valor': indice_prontidao['pct_acesso']
    }, {
        'Indicador': 'Conexão Adequada (≥3 Mbps)',
        'Percentual': f"{indice_prontidao['pct_conexao']:.1f}%",
        'Valor': indice_prontidao['pct_conexao']
    }, {
        'Indicador': 'Formação Docente',
        'Percentual': f"{indice_prontidao['pct_formacao']:.1f}%",
        'Valor': indice_prontidao['pct_formacao']
    }, {
        'Indicador': 'ÍNDICE DE PRONTIDÃO',
        'Percentual': f"{indice_prontidao['indice']:.1f}%",
        'Valor': indice_prontidao['indice']
    }])
    
    df_indice.to_csv(f'{output_dir}/indice_prontidao.csv', index=False, encoding='utf-8-sig')
    print(f"\n✅ Salvo: {output_dir}/indice_prontidao.csv")
    
    # 2. Dados do Paradoxo
    df_paradoxo = pd.DataFrame([{
        'Categoria': 'Alunos que usam IA',
        'Percentual': f"{paradoxo['pct_alunos']:.1f}%",
        'Valor': paradoxo['pct_alunos']
    }, {
        'Categoria': 'Escolas prontas para IA',
        'Percentual': f"{paradoxo['pct_escolas']:.1f}%",
        'Valor': paradoxo['pct_escolas']
    }, {
        'Categoria': 'GAP',
        'Percentual': f"{paradoxo['gap']:.1f} pp",
        'Valor': paradoxo['gap']
    }])
    
    df_paradoxo.to_csv(f'{output_dir}/paradoxo_ia.csv', index=False, encoding='utf-8-sig')
    print(f"✅ Salvo: {output_dir}/paradoxo_ia.csv")
    
    # 3. Dados de uso de IA por alunos
    df_alunos = pd.DataFrame([{
        'Categoria': 'Total Brasil',
        'Usam IA': dados_alunos['brasil']['usa_ia'],
        'Não usam': dados_alunos['brasil']['nao_usa_ia'],
        'Percentual': f"{dados_alunos['brasil']['percentual']:.1f}%"
    }])
    
    df_alunos.to_csv(f'{output_dir}/uso_ia_alunos.csv', index=False, encoding='utf-8-sig')
    print(f"✅ Salvo: {output_dir}/uso_ia_alunos.csv")
    
    print(f"\n📂 Todos os dados salvos em: {output_dir}/")

# ============================================================================
# EXECUÇÃO PRINCIPAL
# ============================================================================

def executar_analise_completa(
    arquivo_escolas_2023='tic_educacao_2023_escolas_tabela_total_v1.0.xlsx',
    arquivo_escolas_2024='tic_educacao_2024_escolas_tabela_total_v1.0.xlsx',
    arquivo_alunos_2024='tic_educacao_2024_alunos_tabela_total_v1.0.xlsx',
    output_dir='./resultados',
    debug=False
):
    """
    Função principal que executa toda a análise
    
    Parâmetros:
    -----------
    arquivo_escolas_2023 : str
        Caminho para o arquivo Excel TIC Educação 2023 - Escolas
    arquivo_escolas_2024 : str
        Caminho para o arquivo Excel TIC Educação 2024 - Escolas
    arquivo_alunos_2024 : str
        Caminho para o arquivo Excel TIC Educação 2024 - Alunos
    output_dir : str
        Diretório onde salvar os resultados
    debug : bool
        Se True, lista todas as abas disponíveis antes de carregar
    """
    print("\n" + "="*80)
    print("ANÁLISE COMPLETA DO TRIPLO DÉFICIT TECNOLÓGICO")
    print("TIC Educação 2023 + 2024")
    print("="*80)
    
    # Modo debug: listar abas disponíveis
    if debug:
        print("\n🔍 MODO DEBUG ATIVADO - Listando abas disponíveis:")
        print("\n" + "-"*80)
        print("ARQUIVO: Escolas 2024")
        listar_abas_disponiveis(arquivo_escolas_2024)
        print("\n" + "-"*80)
        print("ARQUIVO: Alunos 2024")
        listar_abas_disponiveis(arquivo_alunos_2024)
        print("\n" + "-"*80)
    
    # 1. Carregar dados
    print("\n📥 ETAPA 1: CARREGANDO DADOS...")
    dados_escolas_2024 = carregar_dados_escolas_2024(arquivo_escolas_2024)
    dados_alunos_2024 = carregar_dados_alunos_2024(arquivo_alunos_2024)
    
    # 2. Calcular Índice de Prontidão
    print("\n📊 ETAPA 2: CALCULANDO ÍNDICE DE PRONTIDÃO...")
    indice_prontidao = calcular_indice_prontidao(dados_escolas_2024)
    
    # 3. Analisar Paradoxo
    print("\n⚠️  ETAPA 3: ANALISANDO PARADOXO...")
    paradoxo = analisar_paradoxo_ia(dados_alunos_2024, indice_prontidao)
    
    # 4. Criar Visualizações
    print("\n📈 ETAPA 4: GERANDO VISUALIZAÇÕES...")
    fig = criar_visualizacoes_completas(indice_prontidao, paradoxo, dados_alunos_2024)
    
    # Salvar gráfico
    Path(output_dir).mkdir(exist_ok=True)
    fig.savefig(f'{output_dir}/analise_completa_2024.png', dpi=300, bbox_inches='tight')
    print(f"✅ Gráfico salvo: {output_dir}/analise_completa_2024.png")
    
    # 5. Exportar dados
    print("\n💾 ETAPA 5: EXPORTANDO DADOS...")
    exportar_dados_consolidados(indice_prontidao, paradoxo, dados_alunos_2024, output_dir)
    
    # 6. Resumo final
    print("\n" + "="*80)
    print("✅ ANÁLISE CONCLUÍDA COM SUCESSO!")
    print("="*80)
    print(f"\n📊 PRINCIPAIS RESULTADOS:")
    print(f"  • Índice de Prontidão: {indice_prontidao['indice']:.1f}%")
    print(f"  • Alunos que usam IA: {dados_alunos_2024['brasil']['percentual']:.1f}%")
    print(f"  • Gap (Paradoxo): {paradoxo['gap']:.1f} pontos percentuais")
    print(f"\n📂 Arquivos gerados em: {output_dir}/")
    print(f"  • analise_completa_2024.png")
    print(f"  • indice_prontidao.csv")
    print(f"  • paradoxo_ia.csv")
    print(f"  • uso_ia_alunos.csv")
    
    plt.show()
    
    return {
        'indice_prontidao': indice_prontidao,
        'paradoxo': paradoxo,
        'dados_alunos': dados_alunos_2024
    }

# ============================================================================
# EXEMPLO DE USO
# ============================================================================

if __name__ == "__main__":
    """
    Para executar este script:
    
    1. Coloque os arquivos Excel na mesma pasta do script:
       - tic_educacao_2023_escolas_tabela_total_v1.0.xlsx
       - tic_educacao_2024_escolas_tabela_total_v1.0.xlsx
       - tic_educacao_2024_alunos_tabela_total_v1.0.xlsx
    
    2. Execute:
       python edu_br_2024.py
    
    3. Os resultados serão salvos na pasta './resultados/'
    
    MODO DEBUG:
    Se estiver tendo problemas com nomes de abas, use:
       resultados = executar_analise_completa(debug=True)
    
    Isso listará todas as abas disponíveis nos arquivos antes de carregar.
    """
    
    # Executar análise (sem debug)
    resultados = executar_analise_completa()
    
    # Para ativar modo debug, descomente a linha abaixo:
    # resultados = executar_analise_completa(debug=True)
    
    # Acessar resultados individuais se necessário
    print("\n" + "="*80)
    print("DADOS DISPONÍVEIS PARA ANÁLISES ADICIONAIS:")
    print("="*80)
    print("\nresultados['indice_prontidao'] - Dados do índice de prontidão")
    print("resultados['paradoxo'] - Análise do paradoxo uso vs capacidade")
    print("resultados['dados_alunos'] - Dados de uso de IA pelos alunos")